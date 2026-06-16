from dotenv import load_dotenv
load_dotenv()

import os
import re
import json
import uuid
import base64
import logging
import asyncio
from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any

import bcrypt
import jwt
import httpx
import fitz  # PyMuPDF for PDF rasterization
from PIL import Image, ImageOps
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from openpyxl import Workbook
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

# ----------------------------- ENV / DB -----------------------------
ROOT_DIR = Path(__file__).parent
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'change-me')
JWT_ALGORITHM = 'HS256'
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ----------------------------- AUTH HELPERS -----------------------------
ROLES = {"admin", "operations", "finance", "warehouse", "manager"}

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email,
               "exp": datetime.now(timezone.utc) + timedelta(hours=12), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id,
               "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=True,
                        samesite="none", max_age=12*3600, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True,
                        samesite="none", max_age=7*24*3600, path="/")

def clear_auth_cookies(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_roles(*allowed):
    async def checker(user: dict = Depends(get_current_user)):
        if user.get("role") not in allowed:
            raise HTTPException(status_code=403, detail=f"Requires role: {', '.join(allowed)}")
        return user
    return checker

# ----------------------------- MODELS -----------------------------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str
    role: Optional[str] = "operations"
    tenant_name: Optional[str] = None  # if provided, creates a new tenant

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class UserOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    role: str
    tenant_id: str
    tenant_name: Optional[str] = None
    created_at: str

class TenantIn(BaseModel):
    name: str
    gstin: Optional[str] = None

class OcrSettingsIn(BaseModel):
    default_engine: Optional[str] = "gemini"        # gemini | olmocr | auto
    olmocr_endpoint: Optional[str] = ""             # e.g. http://gpu-host:8000
    olmocr_api_key: Optional[str] = ""
    olmocr_model: Optional[str] = "allenai/olmOCR-2-7B-1025-FP8"
    olmocr_timeout: Optional[int] = 120
    auto_fallback_threshold: Optional[float] = 0.5  # if confidence < this in auto mode, fall back
    # Co-Pilot configuration
    copilot_enabled: Optional[bool] = True
    copilot_provider: Optional[str] = "gemini"      # gemini | azure_openai | m365_copilot | gemma | openai | anthropic
    copilot_model_provider: Optional[str] = "gemini"  # legacy alias, kept for backward compat
    copilot_model_name: Optional[str] = "gemini-2.5-pro"
    # Azure OpenAI provider credentials
    azure_endpoint: Optional[str] = ""              # https://<resource>.openai.azure.com
    azure_api_key: Optional[str] = ""
    azure_deployment: Optional[str] = ""            # e.g. gpt-4o-prod
    azure_api_version: Optional[str] = "2024-10-21"
    # Microsoft 365 Copilot provider credentials
    m365_tenant_id: Optional[str] = ""
    m365_client_id: Optional[str] = ""
    m365_client_secret: Optional[str] = ""
    m365_scope: Optional[str] = "https://graph.microsoft.com/.default"
    # Gemma (self-hosted) provider credentials
    gemma_endpoint: Optional[str] = ""              # http://gpu-host:8001 (vLLM/Ollama OpenAI-compatible)
    gemma_api_key: Optional[str] = ""
    gemma_model: Optional[str] = "google/gemma-3-9b-it"
    gemma_timeout: Optional[int] = 60
    # OpenCode Zen (AI gateway, OpenAI-compatible)
    opencode_base_url: Optional[str] = "https://opencode.ai/zen/go/v1"
    opencode_api_key: Optional[str] = ""
    opencode_model: Optional[str] = "deepseek-v4-pro"
    opencode_timeout: Optional[int] = 60

class VendorIn(BaseModel):
    name: str
    gstin: Optional[str] = None
    address: Optional[str] = None
    contact: Optional[str] = None

class DocumentUpdateIn(BaseModel):
    extracted_data: Optional[Dict[str, Any]] = None
    doc_type: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class CopilotChatIn(BaseModel):
    message: str
    history: Optional[List[Dict[str, str]]] = []  # [{role: user|assistant, content: ...}]

class UploadOptionsIn(BaseModel):
    engine_override: Optional[str] = None  # gemini | olmocr | auto

# ----------------------------- VALIDATION -----------------------------
GST_REGEX = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$")

def validate_gst(gstin: Optional[str]) -> bool:
    if not gstin:
        return False
    return bool(GST_REGEX.match(gstin.strip().upper()))

async def run_validations(tenant_id: str, doc_id: str, data: Dict[str, Any]) -> List[Dict[str, str]]:
    errors = []
    mandatory = ["vendor_name", "invoice_number", "invoice_date", "total_amount"]
    for f in mandatory:
        if not data.get(f):
            errors.append({"field": f, "level": "error", "message": f"Missing mandatory field: {f}"})
    gstin = data.get("vendor_gstin") or data.get("gstin")
    if gstin and not validate_gst(gstin):
        errors.append({"field": "vendor_gstin", "level": "warning", "message": "GSTIN format invalid"})
    inv_num = data.get("invoice_number")
    if inv_num:
        dup = await db.documents.find_one({
            "tenant_id": tenant_id,
            "extracted_data.invoice_number": inv_num,
            "id": {"$ne": doc_id}
        }, {"_id": 0, "id": 1})
        if dup:
            errors.append({"field": "invoice_number", "level": "error",
                           "message": f"Duplicate invoice number (also in doc {dup['id'][:8]})"})
    try:
        total = float(data.get("total_amount") or 0)
        items = data.get("line_items") or []
        s = sum(float(i.get("amount") or 0) for i in items)
        if items and abs(total - s) > max(1.0, 0.01 * total):
            errors.append({"field": "total_amount", "level": "warning",
                           "message": f"Total ({total}) does not match line items sum ({s:.2f})"})
    except Exception:
        pass
    return errors

# ----------------------------- AI EXTRACTION -----------------------------
EXTRACTION_PROMPT = """You are an expert document understanding AI for enterprise business documents.

Analyze the attached document image and extract structured data. The document may be one of:
invoice, delivery_challan, purchase_order, grn, packing_slip, eway_bill, transport_slip, receipt, other.

Return ONLY a valid JSON object (no markdown, no explanation) with this schema:
{
  "doc_type": "<one of the above>",
  "confidence": <0.0 to 1.0 overall confidence>,
  "language": "<ISO code, e.g. en, hi>",
  "vendor_name": "string or null",
  "vendor_gstin": "string or null",
  "vendor_address": "string or null",
  "customer_name": "string or null",
  "customer_gstin": "string or null",
  "invoice_number": "string or null",
  "dc_number": "string or null",
  "po_number": "string or null",
  "eway_bill_number": "string or null",
  "invoice_date": "YYYY-MM-DD or null",
  "due_date": "YYYY-MM-DD or null",
  "transport_mode": "string or null",
  "vehicle_number": "string or null",
  "lr_number": "string or null",
  "line_items": [
    {"description": "...", "hsn_sac": "...", "quantity": "...", "unit": "...",
     "unit_price": "...", "amount": "...", "tax_rate": "..."}
  ],
  "subtotal": "number or null",
  "cgst": "number or null",
  "sgst": "number or null",
  "igst": "number or null",
  "total_tax": "number or null",
  "total_amount": "number or null",
  "currency": "string (default INR)",
  "remarks": "string or null",
  "barcode_qr_data": "string or null",
  "raw_text": "<concise plaintext OCR>"
}

If a field is not present, use null. Numeric fields may be returned as numbers or strings.
Be precise: read the document carefully, including handwritten parts, stamps, and signatures."""

async def extract_with_gemini(image_b64: str, mime_type: str) -> Dict[str, Any]:
    if not EMERGENT_LLM_KEY:
        return {"doc_type": "other", "confidence": 0.0, "raw_text": "",
                "error": "EMERGENT_LLM_KEY not configured"}
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"docintel-{uuid.uuid4()}",
            system_message="You are a precise document understanding AI. Always return valid JSON only."
        ).with_model("gemini", "gemini-2.5-pro")
        img = ImageContent(image_base64=image_b64)
        msg = UserMessage(text=EXTRACTION_PROMPT, file_contents=[img])
        response = await chat.send_message(msg)
        text = response.strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?", "", text).rstrip("```").strip()
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            m = re.search(r"\{.*\}", text, re.DOTALL)
            data = json.loads(m.group(0)) if m else {"raw_text": text, "doc_type": "other", "confidence": 0.3}
        data["_engine"] = "gemini"
        return data
    except Exception as e:
        logger.exception("gemini extraction failed")
        return {"doc_type": "other", "confidence": 0.0, "raw_text": "", "error": str(e), "_engine": "gemini"}


def _parse_extraction_text(text: str) -> Dict[str, Any]:
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).rstrip("`").rstrip("```").strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                pass
    return {"raw_text": text, "doc_type": "other", "confidence": 0.3}


async def extract_with_olmocr(image_b64: str, mime_type: str, settings: Dict[str, Any]) -> Dict[str, Any]:
    """Call a user-hosted olmOCR endpoint (OpenAI-compatible, e.g. vLLM-served)."""
    endpoint = (settings.get("olmocr_endpoint") or "").strip().rstrip("/")
    if not endpoint:
        return {"doc_type": "other", "confidence": 0.0, "raw_text": "",
                "error": "olmOCR endpoint not configured", "_engine": "olmocr"}
    api_key = settings.get("olmocr_api_key") or ""
    model = settings.get("olmocr_model") or "allenai/olmOCR-2-7B-1025-FP8"
    timeout = int(settings.get("olmocr_timeout") or 120)
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    payload = {
        "model": model,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "text", "text": EXTRACTION_PROMPT},
                {"type": "image_url",
                 "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}},
            ],
        }],
        "max_tokens": 4096,
        "temperature": 0.0,
    }
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            r = await client.post(f"{endpoint}/v1/chat/completions", json=payload, headers=headers)
            r.raise_for_status()
            result = r.json()
        text = result["choices"][0]["message"]["content"]
        data = _parse_extraction_text(text)
        data["_engine"] = "olmocr"
        data["_engine_model"] = model
        return data
    except httpx.HTTPStatusError as e:
        msg = f"olmOCR HTTP {e.response.status_code}: {e.response.text[:200]}"
        logger.warning(msg)
        return {"doc_type": "other", "confidence": 0.0, "raw_text": "", "error": msg, "_engine": "olmocr"}
    except Exception as e:
        logger.exception("olmocr extraction failed")
        return {"doc_type": "other", "confidence": 0.0, "raw_text": "", "error": str(e), "_engine": "olmocr"}


async def extract_with_engine(tenant_id: str, image_b64: str, mime_type: str,
                              engine_override: Optional[str] = None) -> Dict[str, Any]:
    """(Legacy single-page) Route extraction through tenant-configured engine."""
    return await extract_with_engine_multipage(
        tenant_id,
        [{"file_b64": image_b64, "mime_type": mime_type, "page_number": 1}],
        engine_override)


def _merge_page_extractions(per_page: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Merge N per-page extraction JSONs into a single unified document JSON.
    Headers come from page 1; line_items concat from all pages (tagged with source_page)."""
    if not per_page:
        return {"doc_type": "other", "confidence": 0.0}
    primary = per_page[0]
    out = dict(primary)
    all_items = []
    raw_texts = []
    confidences = []
    for i, p in enumerate(per_page, 1):
        for it in (p.get("line_items") or []):
            all_items.append({**it, "source_page": i})
        if p.get("raw_text"):
            raw_texts.append(f"--- Page {i} ---\n{p['raw_text']}")
        try:
            confidences.append(float(p.get("confidence") or 0))
        except Exception:
            pass
    out["line_items"] = all_items
    if raw_texts:
        out["raw_text"] = "\n\n".join(raw_texts)
    if confidences:
        out["confidence"] = round(sum(confidences) / len(confidences), 3)
    out["_pages_processed"] = len(per_page)
    return out


async def extract_with_engine_multipage(tenant_id: str, pages: List[Dict[str, Any]],
                                        engine_override: Optional[str] = None) -> Dict[str, Any]:
    """Multi-page extraction router.
    Engines with native multi-image support (gemini) → joint call.
    Others (olmocr) → per-page loop + merge."""
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) or {}
    settings = tenant.get("ocr_settings") or {}
    engine = (engine_override or settings.get("default_engine") or "gemini").lower()
    threshold = float(settings.get("auto_fallback_threshold") or 0.5)
    attempts: List[Dict[str, Any]] = []

    async def _gemini_joint() -> Dict[str, Any]:
        if not EMERGENT_LLM_KEY:
            return {"doc_type": "other", "confidence": 0.0, "raw_text": "",
                    "error": "EMERGENT_LLM_KEY not configured", "_engine": "gemini"}
        try:
            chat = LlmChat(
                api_key=EMERGENT_LLM_KEY,
                session_id=f"docintel-{uuid.uuid4()}",
                system_message="You are a precise document understanding AI. Always return valid JSON only."
            ).with_model("gemini", "gemini-2.5-pro")
            imgs = [ImageContent(image_base64=p["file_b64"]) for p in pages]
            prompt = EXTRACTION_PROMPT + (
                f"\n\nThis document has {len(pages)} pages provided in order. "
                "Combine information across pages — header on page 1, line items may span pages."
                if len(pages) > 1 else "")
            response = await chat.send_message(UserMessage(text=prompt, file_contents=imgs))
            data = _parse_extraction_text(response)
            data["_engine"] = "gemini"
            data["_pages_processed"] = len(pages)
            return data
        except Exception as e:
            logger.exception("gemini joint extraction failed")
            return {"doc_type": "other", "confidence": 0.0, "raw_text": "", "error": str(e), "_engine": "gemini"}

    async def _olmocr_per_page() -> Dict[str, Any]:
        results = []
        any_error = None
        for p in pages:
            r = await extract_with_olmocr(p["file_b64"], p["mime_type"], settings)
            if r.get("error"):
                any_error = r.get("error")
            results.append(r)
        merged = _merge_page_extractions(results)
        merged["_engine"] = "olmocr"
        if any_error and all(r.get("error") for r in results):
            merged["error"] = any_error
        return merged

    if engine == "olmocr":
        data = await _olmocr_per_page()
        attempts.append({"engine": "olmocr", "ok": not data.get("error"),
                         "confidence": data.get("confidence", 0)})
        data["_attempts"] = attempts
        return data

    if engine == "gemini":
        data = await _gemini_joint()
        attempts.append({"engine": "gemini", "ok": not data.get("error"),
                         "confidence": data.get("confidence", 0)})
        data["_attempts"] = attempts
        return data

    # auto
    if settings.get("olmocr_endpoint"):
        primary = await _olmocr_per_page()
        ok = not primary.get("error") and float(primary.get("confidence") or 0) >= threshold
        attempts.append({"engine": "olmocr", "ok": ok,
                         "confidence": primary.get("confidence", 0),
                         "error": primary.get("error")})
        if ok:
            primary["_attempts"] = attempts
            return primary
    fallback = await _gemini_joint()
    attempts.append({"engine": "gemini", "ok": not fallback.get("error"),
                     "confidence": fallback.get("confidence", 0)})
    fallback["_attempts"] = attempts
    fallback["_engine"] = "gemini"
    return fallback


# ----------------------------- COPILOT CHAT -----------------------------
COPILOT_SYSTEM = """You are DocIntel Co-Pilot — an AI assistant embedded in an enterprise document
intelligence platform. The user is reviewing an already-extracted business document (invoice, DC,
GRN, etc.). You will be given:
- the source document image
- the currently extracted structured fields (JSON)
- any pending validation errors
- the user's question or instruction

Be concise, accurate, and helpful. When the user asks for a correction, return exactly what to change
(field name + new value). When asked to explain a discrepancy, cite specific values. Avoid hallucinating
data not visible in the document. Reply in plain text or short markdown."""


def _build_copilot_context(doc: dict) -> str:
    return (
        f"Document filename: {doc.get('filename')}\n"
        f"Detected type: {doc.get('doc_type')}\n"
        f"Confidence: {doc.get('confidence')}\n"
        f"Engine used: {(doc.get('extracted_data') or {}).get('_engine')}\n"
        f"Extracted fields (JSON):\n{json.dumps(doc.get('extracted_data') or {}, indent=2, default=str)}\n\n"
        f"Validation errors:\n{json.dumps(doc.get('validation_errors') or [], indent=2)}\n"
    )


async def _copilot_emergent(provider: str, model: str, system_msg: str, user_text: str,
                            file_b64: str, mime: str, doc_id: str) -> str:
    """Use emergentintegrations (works for gemini/openai/anthropic via Emergent LLM Key)."""
    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"copilot-{doc_id}-{uuid.uuid4()}",
        system_message=system_msg,
    ).with_model(provider, model)
    attachments = []
    if mime.startswith("image/") and file_b64:
        attachments.append(ImageContent(image_base64=file_b64))
    return (await chat.send_message(UserMessage(text=user_text, file_contents=attachments))).strip()


async def _copilot_azure_openai(settings: Dict[str, Any], system_msg: str, user_text: str,
                                history: List[Dict[str, str]], file_b64: str, mime: str) -> str:
    endpoint = (settings.get("azure_endpoint") or "").strip().rstrip("/")
    api_key = settings.get("azure_api_key") or ""
    deployment = settings.get("azure_deployment") or ""
    api_version = settings.get("azure_api_version") or "2024-10-21"
    if not (endpoint and api_key and deployment):
        return "Azure OpenAI is not fully configured (need endpoint, api_key, deployment)."
    msgs = [{"role": "system", "content": system_msg}]
    for m in (history or [])[-8:]:
        msgs.append({"role": m.get("role", "user"), "content": m.get("content", "")})
    if mime.startswith("image/") and file_b64:
        msgs.append({"role": "user", "content": [
            {"type": "text", "text": user_text},
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{file_b64}"}},
        ]})
    else:
        msgs.append({"role": "user", "content": user_text})
    url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?api-version={api_version}"
    try:
        async with httpx.AsyncClient(timeout=45) as client:
            r = await client.post(url, headers={"api-key": api_key, "Content-Type": "application/json"},
                                  json={"messages": msgs, "max_tokens": 1024, "temperature": 0.3})
            r.raise_for_status()
            data = r.json()
        return (data.get("choices") or [{}])[0].get("message", {}).get("content", "").strip() or "(empty reply)"
    except httpx.HTTPStatusError as e:
        return f"Azure OpenAI error {e.response.status_code}: {e.response.text[:200]}"
    except Exception as e:
        return f"Azure OpenAI request failed: {e}"


async def _m365_get_token(tenant_id_db: str, settings: Dict[str, Any]) -> Optional[str]:
    """Get M365 access token via client_credentials, cached in MongoDB."""
    m_tid = (settings.get("m365_tenant_id") or "").strip()
    cid = (settings.get("m365_client_id") or "").strip()
    secret = (settings.get("m365_client_secret") or "").strip()
    scope = settings.get("m365_scope") or "https://graph.microsoft.com/.default"
    if not (m_tid and cid and secret):
        return None
    cache = await db.token_cache.find_one({"tenant_id": tenant_id_db, "client_id": cid})
    if cache and cache.get("expires_at"):
        try:
            exp = datetime.fromisoformat(cache["expires_at"])
            if exp > datetime.now(timezone.utc) + timedelta(minutes=2):
                return cache["access_token"]
        except Exception:
            pass
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.post(
                f"https://login.microsoftonline.com/{m_tid}/oauth2/v2.0/token",
                data={"grant_type": "client_credentials", "client_id": cid,
                      "client_secret": secret, "scope": scope})
            r.raise_for_status()
            tok = r.json()
        await db.token_cache.update_one(
            {"tenant_id": tenant_id_db, "client_id": cid},
            {"$set": {"access_token": tok["access_token"],
                      "expires_at": (datetime.now(timezone.utc) + timedelta(seconds=tok["expires_in"])).isoformat(),
                      "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True)
        return tok["access_token"]
    except Exception as e:
        logger.warning("m365 token failed: %s", e)
        return None


async def _copilot_m365(tenant_id_db: str, settings: Dict[str, Any], user_text: str,
                        history: List[Dict[str, str]], context: str) -> str:
    token = await _m365_get_token(tenant_id_db, settings)
    if not token:
        return ("Microsoft 365 Copilot is not fully configured (need tenant_id, client_id, client_secret) "
                "or token acquisition failed.")
    # Compose grounding context as the first user message
    convo_msgs = [{"role": "user", "content": f"CONTEXT:\n{context}"}]
    for m in (history or [])[-8:]:
        convo_msgs.append({"role": m.get("role", "user"), "content": m.get("content", "")})
    convo_msgs.append({"role": "user", "content": user_text})
    try:
        async with httpx.AsyncClient(timeout=45) as client:
            # 1) create a conversation
            cr = await client.post(
                "https://graph.microsoft.com/beta/copilot/conversations",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={})
            if cr.status_code >= 400:
                return f"M365 Copilot conversation create failed ({cr.status_code}): {cr.text[:200]}"
            conv_id = cr.json().get("id")
            # 2) send chat
            sr = await client.post(
                f"https://graph.microsoft.com/beta/copilot/conversations/{conv_id}/chat",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={"messages": convo_msgs})
            if sr.status_code >= 400:
                return (f"M365 Copilot chat failed ({sr.status_code}): {sr.text[:300]}. "
                        f"Note: the M365 Copilot Chat API requires delegated permissions and admin consent; "
                        f"application-only tokens may be rejected.")
            data = sr.json()
        # Response shape varies; try common paths
        if isinstance(data, dict):
            msgs = data.get("messages") or data.get("value") or []
            for m in reversed(msgs):
                c = m.get("content") if isinstance(m, dict) else None
                if c:
                    return c if isinstance(c, str) else json.dumps(c)
            if data.get("text"):
                return str(data["text"])
        return json.dumps(data)[:600]
    except Exception as e:
        logger.exception("m365 copilot failed")
        return f"M365 Copilot request failed: {e}"


async def _copilot_gemma(settings: Dict[str, Any], system_msg: str, user_text: str,
                         history: List[Dict[str, str]], file_b64: str, mime: str) -> str:
    """Call a self-hosted Gemma instance (vLLM / Ollama OpenAI-compatible API)."""
    endpoint = (settings.get("gemma_endpoint") or "").strip().rstrip("/")
    if not endpoint:
        return "Gemma endpoint is not configured."
    api_key = settings.get("gemma_api_key") or ""
    model = settings.get("gemma_model") or "google/gemma-3-9b-it"
    timeout = int(settings.get("gemma_timeout") or 60)
    msgs = [{"role": "system", "content": system_msg}]
    for m in (history or [])[-8:]:
        msgs.append({"role": m.get("role", "user"), "content": m.get("content", "")})
    if mime.startswith("image/") and file_b64:
        msgs.append({"role": "user", "content": [
            {"type": "text", "text": user_text},
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{file_b64}"}},
        ]})
    else:
        msgs.append({"role": "user", "content": user_text})
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            r = await client.post(f"{endpoint}/v1/chat/completions",
                                  headers=headers,
                                  json={"model": model, "messages": msgs, "max_tokens": 1024, "temperature": 0.3})
            r.raise_for_status()
            data = r.json()
        return (data.get("choices") or [{}])[0].get("message", {}).get("content", "").strip() or "(empty reply)"
    except httpx.HTTPStatusError as e:
        return f"Gemma error {e.response.status_code}: {e.response.text[:200]}"
    except Exception as e:
        return f"Gemma request failed: {e}"


async def _copilot_opencode(settings: Dict[str, Any], system_msg: str, user_text: str,
                            history: List[Dict[str, str]], file_b64: str, mime: str) -> str:
    """Call OpenCode Zen (AI gateway, OpenAI-compatible chat completions)."""
    base = (settings.get("opencode_base_url") or "https://opencode.ai/zen/go/v1").strip().rstrip("/")
    api_key = settings.get("opencode_api_key") or ""
    model = settings.get("opencode_model") or "deepseek-v4-pro"
    timeout = int(settings.get("opencode_timeout") or 60)
    if not api_key:
        return "OpenCode Zen API key is not configured. Add it in Settings → Co-Pilot."
    # Whitelist of known multimodal models in OpenCode Zen catalogue (best-effort)
    vision_capable = {"mimo-v2-omni", "mimo-v2.5-pro", "minimax-m3"}
    use_image = mime.startswith("image/") and file_b64 and model in vision_capable
    msgs = [{"role": "system", "content": system_msg}]
    for m in (history or [])[-8:]:
        msgs.append({"role": m.get("role", "user"), "content": m.get("content", "")})
    if use_image:
        msgs.append({"role": "user", "content": [
            {"type": "text", "text": user_text},
            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{file_b64}"}},
        ]})
    else:
        msgs.append({"role": "user", "content": user_text})
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            r = await client.post(f"{base}/chat/completions",
                                  headers=headers,
                                  json={"model": model, "messages": msgs, "max_tokens": 1024, "temperature": 0.3})
            r.raise_for_status()
            data = r.json()
        return (data.get("choices") or [{}])[0].get("message", {}).get("content", "").strip() or "(empty reply)"
    except httpx.HTTPStatusError as e:
        return f"OpenCode Zen error {e.response.status_code}: {e.response.text[:200]}"
    except Exception as e:
        return f"OpenCode Zen request failed: {e}"


async def copilot_reply(tenant_id: str, doc: dict, file_b64: str, mime: str,
                        user_msg: str, history: List[Dict[str, str]]) -> str:
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) or {}
    settings = tenant.get("ocr_settings") or {}
    if settings.get("copilot_enabled") is False:
        return "Co-Pilot is disabled for this tenant. Ask your admin to enable it in Settings."

    provider = (settings.get("copilot_provider")
                or settings.get("copilot_model_provider")
                or "gemini").lower()
    model = settings.get("copilot_model_name") or "gemini-2.5-pro"

    context = _build_copilot_context(doc)
    convo_text = ""
    for m in (history or [])[-8:]:
        convo_text += f"\n[{m.get('role','user').upper()}]: {m.get('content','')}\n"
    full_text = f"{context}\n\n--- Conversation so far ---{convo_text}\n[USER]: {user_msg}"

    try:
        if provider in ("gemini", "openai", "anthropic"):
            return await _copilot_emergent(provider, model, COPILOT_SYSTEM, full_text,
                                           file_b64, mime, doc.get("id", ""))
        if provider == "azure_openai":
            return await _copilot_azure_openai(settings, COPILOT_SYSTEM, user_msg,
                                                [{"role": "user", "content": f"CONTEXT:\n{context}"}] + (history or []),
                                                file_b64, mime)
        if provider == "m365_copilot":
            return await _copilot_m365(tenant_id, settings, user_msg, history or [], context)
        if provider == "gemma":
            return await _copilot_gemma(settings, COPILOT_SYSTEM, user_msg,
                                         [{"role": "user", "content": f"CONTEXT:\n{context}"}] + (history or []),
                                         file_b64, mime)
        if provider == "opencode_zen":
            return await _copilot_opencode(settings, COPILOT_SYSTEM, user_msg,
                                            [{"role": "user", "content": f"CONTEXT:\n{context}"}] + (history or []),
                                            file_b64, mime)
        return f"Unknown copilot provider: {provider}"
    except Exception as e:
        logger.exception("copilot failed")
        return f"Co-Pilot error: {e}"

# ----------------------------- APP / ROUTER -----------------------------
app = FastAPI(title="Document Intelligence Platform API", version="1.0.0",
              description="Enterprise AI-powered Document Intelligence Platform")
api = APIRouter(prefix="/api")

# ----------------------------- AUTH ROUTES -----------------------------
async def log_audit(tenant_id: str, user_id: str, action: str, resource_type: str,
                    resource_id: str = "", details: Optional[Dict] = None):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "user_id": user_id,
        "action": action,
        "resource_type": resource_type,
        "resource_id": resource_id,
        "details": details or {},
        "created_at": datetime.now(timezone.utc).isoformat()
    })

@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower().strip()
    # Public registration ALWAYS creates non-admin user. Admins are elevated via /users by an existing admin.
    requested_role = (payload.role or "operations").lower()
    role = requested_role if requested_role in (ROLES - {"admin"}) else "operations"
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already registered")
    # If tenant_name provided, create new tenant; otherwise default tenant
    if payload.tenant_name:
        tenant_id = str(uuid.uuid4())
        await db.tenants.insert_one({
            "id": tenant_id, "name": payload.tenant_name, "gstin": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    else:
        default = await db.tenants.find_one({"name": "Default Tenant"}, {"_id": 0})
        if not default:
            tenant_id = str(uuid.uuid4())
            await db.tenants.insert_one({
                "id": tenant_id, "name": "Default Tenant", "gstin": None,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        else:
            tenant_id = default["id"]
    uid = str(uuid.uuid4())
    user_doc = {
        "id": uid, "email": email, "password_hash": hash_password(payload.password),
        "name": payload.name, "role": role, "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)
    access = create_access_token(uid, email)
    refresh = create_refresh_token(uid)
    set_auth_cookies(response, access, refresh)
    user_doc.pop("password_hash", None)
    user_doc.pop("_id", None)
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    user_doc["tenant_name"] = tenant.get("name") if tenant else None
    await log_audit(tenant_id, uid, "register", "user", uid)
    return user_doc

@api.post("/auth/login")
async def login(payload: LoginIn, request: Request, response: Response):
    email = payload.email.lower().strip()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        last = datetime.fromisoformat(attempt["last"])
        if (datetime.now(timezone.utc) - last).total_seconds() < 900:
            raise HTTPException(429, "Too many attempts. Try again in 15 minutes.")
        await db.login_attempts.delete_one({"identifier": identifier})
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        raise HTTPException(401, "Invalid email or password")
    await db.login_attempts.delete_one({"identifier": identifier})
    access = create_access_token(user["id"], email)
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    user.pop("password_hash", None)
    user.pop("_id", None)
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    user["tenant_name"] = tenant.get("name") if tenant else None
    await log_audit(user["tenant_id"], user["id"], "login", "user", user["id"])
    return user

@api.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    clear_auth_cookies(response)
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    tenant = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    user["tenant_name"] = tenant.get("name") if tenant else None
    return user

@api.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(401, "No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(401, "Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(401, "User not found")
        access = create_access_token(user["id"], user["email"])
        response.set_cookie("access_token", access, httponly=True, secure=True,
                            samesite="none", max_age=12*3600, path="/")
        return {"ok": True}
    except jwt.PyJWTError:
        raise HTTPException(401, "Invalid refresh token")

# ----------------------------- TENANTS / USERS -----------------------------
@api.get("/tenants/me")
async def get_my_tenant(user: dict = Depends(get_current_user)):
    t = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})
    return t or {}

@api.put("/tenants/me")
async def update_my_tenant(payload: TenantIn, user: dict = Depends(require_roles("admin"))):
    await db.tenants.update_one({"id": user["tenant_id"]},
                                {"$set": {"name": payload.name, "gstin": payload.gstin}})
    return await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0})

SECRET_FIELDS = ["olmocr_api_key", "azure_api_key", "m365_client_secret", "gemma_api_key", "opencode_api_key"]

def _mask_secret(v: Optional[str]) -> str:
    if not v:
        return ""
    if str(v).startswith("***"):
        return str(v)
    s = str(v)
    return "***" + s[-4:] if len(s) >= 4 else "***"


@api.get("/tenants/me/ocr-settings")
async def get_ocr_settings(user: dict = Depends(get_current_user)):
    t = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0}) or {}
    s = dict(t.get("ocr_settings") or {})
    # mask all secret fields before returning
    for f in SECRET_FIELDS:
        if s.get(f):
            s[f] = _mask_secret(s[f])
    defaults = {
        "default_engine": "gemini",
        "olmocr_endpoint": "",
        "olmocr_api_key": "",
        "olmocr_model": "allenai/olmOCR-2-7B-1025-FP8",
        "olmocr_timeout": 120,
        "auto_fallback_threshold": 0.5,
        "copilot_enabled": True,
        "copilot_provider": "gemini",
        "copilot_model_provider": "gemini",
        "copilot_model_name": "gemini-2.5-pro",
        "azure_endpoint": "",
        "azure_api_key": "",
        "azure_deployment": "",
        "azure_api_version": "2024-10-21",
        "m365_tenant_id": "",
        "m365_client_id": "",
        "m365_client_secret": "",
        "m365_scope": "https://graph.microsoft.com/.default",
        "gemma_endpoint": "",
        "gemma_api_key": "",
        "gemma_model": "google/gemma-3-9b-it",
        "gemma_timeout": 60,
        "opencode_base_url": "https://opencode.ai/zen/go/v1",
        "opencode_api_key": "",
        "opencode_model": "deepseek-v4-pro",
        "opencode_timeout": 60,
    }
    return {**defaults, **s}

@api.put("/tenants/me/ocr-settings")
async def update_ocr_settings(payload: OcrSettingsIn, user: dict = Depends(require_roles("admin"))):
    # Only fields explicitly sent by the client should overwrite existing values
    incoming = payload.model_dump(exclude_unset=True)
    engine = (incoming.get("default_engine") or "").lower() if "default_engine" in incoming else None
    if engine and engine not in ("gemini", "olmocr", "auto"):
        raise HTTPException(400, "default_engine must be one of: gemini, olmocr, auto")
    provider = (incoming.get("copilot_provider") or "").lower() if "copilot_provider" in incoming else None
    if provider and provider not in ("gemini", "openai", "anthropic", "azure_openai", "m365_copilot", "gemma", "opencode_zen"):
        raise HTTPException(400, "copilot_provider must be one of: gemini, openai, anthropic, azure_openai, m365_copilot, gemma, opencode_zen")
    # mirror copilot_provider to legacy copilot_model_provider for back-compat consumers
    if provider:
        incoming["copilot_model_provider"] = provider
    # if any secret comes in as masked value, do not overwrite the real one
    for f in SECRET_FIELDS:
        if str(incoming.get(f, "")).startswith("***"):
            incoming.pop(f, None)
    existing = (await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0}) or {}).get("ocr_settings") or {}
    merged = {**existing, **incoming}
    await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": {"ocr_settings": merged}})
    await log_audit(user["tenant_id"], user["id"], "update_ocr_settings", "tenant",
                    user["tenant_id"], {"engine": merged.get("default_engine"),
                                        "copilot_provider": merged.get("copilot_provider")})
    masked = dict(merged)
    for f in SECRET_FIELDS:
        if masked.get(f):
            masked[f] = _mask_secret(masked[f])
    return masked

@api.post("/tenants/me/ocr-settings/test")
async def test_olmocr(user: dict = Depends(require_roles("admin"))):
    """Ping the configured olmOCR endpoint to verify connectivity (no image sent)."""
    t = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0}) or {}
    s = t.get("ocr_settings") or {}
    endpoint = (s.get("olmocr_endpoint") or "").strip().rstrip("/")
    if not endpoint:
        raise HTTPException(400, "olmocr_endpoint not configured")
    headers = {}
    if s.get("olmocr_api_key"):
        headers["Authorization"] = f"Bearer {s['olmocr_api_key']}"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"{endpoint}/v1/models", headers=headers)
        return {"ok": r.status_code == 200, "status": r.status_code,
                "body": r.text[:400]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# 5-minute in-memory cache for OpenCode Zen model catalogue (keyed by tenant_id)
_OPENCODE_MODEL_CACHE: Dict[str, Dict[str, Any]] = {}

@api.get("/tenants/me/copilot/models")
async def list_copilot_models(provider: str = Query("opencode_zen"),
                              refresh: bool = Query(False),
                              user: dict = Depends(get_current_user)):
    """Fetch the live model catalogue for a Co-Pilot provider that supports /models discovery."""
    if provider != "opencode_zen":
        # For now, only OpenCode Zen supports dynamic model discovery here.
        return {"provider": provider, "models": [], "note": "Model discovery not implemented for this provider."}
    t = await db.tenants.find_one({"id": user["tenant_id"]}, {"_id": 0}) or {}
    s = t.get("ocr_settings") or {}
    base = (s.get("opencode_base_url") or "https://opencode.ai/zen/go/v1").strip().rstrip("/")
    api_key = s.get("opencode_api_key") or ""
    cache_key = f"{user['tenant_id']}::{base}"
    now = datetime.now(timezone.utc)
    cached = _OPENCODE_MODEL_CACHE.get(cache_key)
    if cached and not refresh:
        if (now - cached["fetched_at"]).total_seconds() < 300:
            return {"provider": "opencode_zen", "models": cached["models"],
                    "cached": True, "fetched_at": cached["fetched_at"].isoformat()}
    headers = {}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"{base}/models", headers=headers)
            r.raise_for_status()
            payload = r.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(e.response.status_code,
                            f"OpenCode Zen error: {e.response.text[:200]}")
    except Exception as e:
        raise HTTPException(502, f"OpenCode Zen request failed: {e}")
    raw = payload.get("data") or []
    # known multimodal hints
    vision_capable = {"mimo-v2-omni", "mimo-v2.5-pro", "minimax-m3"}
    models = [{
        "id": m.get("id"),
        "owned_by": m.get("owned_by"),
        "created": m.get("created"),
        "multimodal": m.get("id") in vision_capable,
    } for m in raw if m.get("id")]
    _OPENCODE_MODEL_CACHE[cache_key] = {"models": models, "fetched_at": now}
    return {"provider": "opencode_zen", "models": models,
            "cached": False, "fetched_at": now.isoformat()}

@api.get("/users")
async def list_users(user: dict = Depends(require_roles("admin", "manager"))):
    cursor = db.users.find({"tenant_id": user["tenant_id"]}, {"_id": 0, "password_hash": 0})
    return await cursor.to_list(500)

@api.post("/users")
async def create_user(payload: RegisterIn, user: dict = Depends(require_roles("admin"))):
    email = payload.email.lower().strip()
    role = (payload.role or "operations").lower()
    if role not in ROLES:
        raise HTTPException(400, "Invalid role")
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already exists")
    uid = str(uuid.uuid4())
    doc = {"id": uid, "email": email, "password_hash": hash_password(payload.password),
           "name": payload.name, "role": role, "tenant_id": user["tenant_id"],
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.users.insert_one(doc)
    await log_audit(user["tenant_id"], user["id"], "create_user", "user", uid)
    doc.pop("password_hash", None)
    doc.pop("_id", None)
    return doc

@api.delete("/users/{uid}")
async def delete_user(uid: str, user: dict = Depends(require_roles("admin"))):
    if uid == user["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    res = await db.users.delete_one({"id": uid, "tenant_id": user["tenant_id"]})
    if res.deleted_count == 0:
        raise HTTPException(404, "Not found")
    await log_audit(user["tenant_id"], user["id"], "delete_user", "user", uid)
    return {"ok": True}

# ----------------------------- VENDORS -----------------------------
@api.get("/vendors")
async def list_vendors(user: dict = Depends(get_current_user)):
    cursor = db.vendors.find({"tenant_id": user["tenant_id"]}, {"_id": 0})
    return await cursor.to_list(1000)

@api.post("/vendors")
async def create_vendor(payload: VendorIn, user: dict = Depends(require_roles("admin", "operations", "finance"))):
    if payload.gstin and not validate_gst(payload.gstin):
        raise HTTPException(400, "Invalid GSTIN format")
    vid = str(uuid.uuid4())
    doc = {"id": vid, "tenant_id": user["tenant_id"], **payload.model_dump(),
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.vendors.insert_one(doc)
    doc.pop("_id", None)
    await log_audit(user["tenant_id"], user["id"], "create_vendor", "vendor", vid)
    return doc

@api.delete("/vendors/{vid}")
async def delete_vendor(vid: str, user: dict = Depends(require_roles("admin", "operations"))):
    res = await db.vendors.delete_one({"id": vid, "tenant_id": user["tenant_id"]})
    if res.deleted_count == 0:
        raise HTTPException(404, "Not found")
    await log_audit(user["tenant_id"], user["id"], "delete_vendor", "vendor", vid)
    return {"ok": True}

# ----------------------------- DOCUMENTS -----------------------------
# ----------------------------- IMAGE & PDF HELPERS -----------------------------
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "30"))
IMAGE_MAX_DIM = int(os.environ.get("IMAGE_MAX_DIM", "2048"))
IMAGE_JPEG_QUALITY = int(os.environ.get("IMAGE_JPEG_QUALITY", "85"))
COMPRESS_THRESHOLD_BYTES = int(os.environ.get("COMPRESS_THRESHOLD_BYTES", str(1024 * 1024)))
MAX_PAGES_PER_DOC = int(os.environ.get("MAX_PAGES_PER_DOC", "3"))
PDF_RENDER_DPI = int(os.environ.get("PDF_RENDER_DPI", "200"))


def _compress_image_bytes(content: bytes, mime: str) -> tuple[bytes, str, int, int]:
    """Auto-rotate via EXIF, resize to IMAGE_MAX_DIM long edge, JPEG q=85.
    Returns (bytes, new_mime, width, height). No-op if already small enough."""
    try:
        with Image.open(BytesIO(content)) as im:
            im = ImageOps.exif_transpose(im)
            w, h = im.size
            long_edge = max(w, h)
            needs_resize = long_edge > IMAGE_MAX_DIM
            needs_recompress = len(content) > COMPRESS_THRESHOLD_BYTES or mime.lower() == "image/png"
            if not (needs_resize or needs_recompress):
                return content, mime, w, h
            if needs_resize:
                scale = IMAGE_MAX_DIM / long_edge
                im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
            if im.mode in ("RGBA", "P", "LA"):
                bg = Image.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[-1] if im.mode in ("RGBA", "LA") else None)
                im = bg
            elif im.mode != "RGB":
                im = im.convert("RGB")
            out = BytesIO()
            im.save(out, format="JPEG", quality=IMAGE_JPEG_QUALITY, optimize=True)
            return out.getvalue(), "image/jpeg", im.width, im.height
    except Exception as e:
        logger.warning("image compress failed (%s); using original", e)
        return content, mime, 0, 0


def _rasterize_pdf(content: bytes) -> list[dict]:
    """Rasterize each PDF page to JPEG. Returns list of page dicts (capped at MAX_PAGES_PER_DOC)."""
    pages = []
    try:
        doc = fitz.open(stream=content, filetype="pdf")
        zoom = PDF_RENDER_DPI / 72.0
        mat = fitz.Matrix(zoom, zoom)
        for i, page in enumerate(doc):
            if i >= MAX_PAGES_PER_DOC:
                break
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("png")
            compressed, mime, w, h = _compress_image_bytes(img_bytes, "image/png")
            pages.append({
                "page_number": i + 1,
                "mime_type": mime,
                "file_b64": base64.b64encode(compressed).decode("utf-8"),
                "width": w, "height": h,
                "size": len(compressed),
            })
        doc.close()
    except Exception as e:
        logger.exception("pdf rasterize failed")
        raise HTTPException(400, f"Could not read PDF: {e}")
    if not pages:
        raise HTTPException(400, "PDF has no rasterizable pages")
    return pages


def _build_pages_from_files(files: list[UploadFile], contents: list[bytes]) -> list[dict]:
    """Combine multiple file uploads (images and/or PDFs) into a single pages[] list."""
    out: list[dict] = []
    for upload, content in zip(files, contents):
        mime = upload.content_type or "application/octet-stream"
        if mime == "application/octet-stream":
            ext = (upload.filename or "").lower().rsplit(".", 1)[-1]
            mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "webp": "image/webp", "pdf": "application/pdf"}.get(ext, mime)
        if mime == "application/pdf":
            out.extend(_rasterize_pdf(content))
        elif mime.startswith("image/"):
            compressed, new_mime, w, h = _compress_image_bytes(content, mime)
            out.append({
                "page_number": len(out) + 1,
                "mime_type": new_mime,
                "file_b64": base64.b64encode(compressed).decode("utf-8"),
                "width": w, "height": h,
                "size": len(compressed),
            })
        else:
            raise HTTPException(400, f"Unsupported file type: {mime}")
        if len(out) >= MAX_PAGES_PER_DOC:
            break
    out = out[:MAX_PAGES_PER_DOC]
    for idx, p in enumerate(out, 1):
        p["page_number"] = idx
    return out


def _doc_pages(doc: dict) -> list[dict]:
    """Return pages[] from a document, synthesising one from legacy file_b64 if needed."""
    pages = doc.get("pages")
    if pages:
        return pages
    if doc.get("file_b64"):
        return [{
            "page_number": 1,
            "mime_type": doc.get("mime_type") or "image/jpeg",
            "file_b64": doc["file_b64"],
            "width": 0, "height": 0,
            "size": doc.get("size", 0),
        }]
    return []


# ----------------------------- VISIBILITY -----------------------------
def _visibility_filter(user: dict) -> dict:
    """Row-level scope on documents.
    admin/finance/manager → entire tenant; operations/warehouse → only own uploads."""
    base = {"tenant_id": user["tenant_id"]}
    if user.get("role") in ("admin", "finance", "manager"):
        return base
    return {**base, "uploaded_by": user["id"]}


async def _save_document(file: UploadFile, user: dict) -> dict:
    content = await file.read()
    if len(content) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(413, f"File too large: {file.filename}")
    pages = _build_pages_from_files([file], [content])
    first = pages[0]
    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": user["tenant_id"],
        "filename": file.filename,
        # legacy single-page fields (first page) for backward compat
        "mime_type": first["mime_type"],
        "file_b64": first["file_b64"],
        "size": sum(p.get("size", 0) for p in pages),
        # new multi-page fields
        "pages": pages,
        "page_count": len(pages),
        "doc_type": "unknown",
        "status": "pending",
        "extracted_data": {},
        "confidence": 0.0,
        "validation_errors": [],
        "uploaded_by": user["id"],
        "uploaded_by_name": user.get("name"),
        "approved_by": None,
        "notes": "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.documents.insert_one(doc.copy())
    await log_audit(user["tenant_id"], user["id"], "upload_document", "document", doc["id"],
                    {"filename": file.filename, "pages": len(pages), "size": doc["size"]})
    return doc


async def _save_document_multipage(files: list[UploadFile], user: dict, group_filename: str = None) -> dict:
    """Save multiple uploaded files as a SINGLE multi-page document."""
    contents = []
    total_size = 0
    for f in files:
        c = await f.read()
        total_size += len(c)
        if total_size > MAX_UPLOAD_MB * 1024 * 1024:
            raise HTTPException(413, "Combined files too large")
        contents.append(c)
    pages = _build_pages_from_files(files, contents)
    first = pages[0]
    fname = group_filename or (files[0].filename if files else "document")
    if len(files) > 1 and not group_filename:
        fname = f"{fname} (+{len(files)-1} more)"
    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": user["tenant_id"],
        "filename": fname,
        "mime_type": first["mime_type"],
        "file_b64": first["file_b64"],
        "size": sum(p.get("size", 0) for p in pages),
        "pages": pages,
        "page_count": len(pages),
        "doc_type": "unknown",
        "status": "pending",
        "extracted_data": {},
        "confidence": 0.0,
        "validation_errors": [],
        "uploaded_by": user["id"],
        "uploaded_by_name": user.get("name"),
        "approved_by": None,
        "notes": "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.documents.insert_one(doc.copy())
    await log_audit(user["tenant_id"], user["id"], "upload_document", "document", doc["id"],
                    {"filename": fname, "pages": len(pages), "size": doc["size"], "multipage": True})
    return doc

@api.post("/documents/upload")
async def upload_document(file: UploadFile = File(...),
                          auto_process: bool = Form(True),
                          engine_override: Optional[str] = Form(None),
                          user: dict = Depends(get_current_user)):
    doc = await _save_document(file, user)
    if auto_process:
        asyncio.create_task(_process_document_async(doc["id"], user["tenant_id"], engine_override))
    doc.pop("file_b64", None)
    for p in doc.get("pages") or []:
        p.pop("file_b64", None)
    return doc

@api.post("/documents/upload-bulk")
async def upload_bulk(files: List[UploadFile] = File(...),
                      auto_process: bool = Form(True),
                      engine_override: Optional[str] = Form(None),
                      as_single_document: bool = Form(False),
                      group_filename: Optional[str] = Form(None),
                      user: dict = Depends(get_current_user)):
    """If as_single_document=True, combine ALL files into one multi-page document.
    Otherwise each file becomes its own document (default, existing behaviour)."""
    if as_single_document:
        try:
            d = await _save_document_multipage(files, user, group_filename)
            if auto_process:
                asyncio.create_task(_process_document_async(d["id"], user["tenant_id"], engine_override))
            d.pop("file_b64", None)
            for p in d.get("pages") or []:
                p.pop("file_b64", None)
            return {"uploaded": 1, "documents": [d]}
        except HTTPException as e:
            return {"uploaded": 0, "documents": [{"error": e.detail}]}
    saved = []
    for f in files:
        try:
            d = await _save_document(f, user)
            if auto_process:
                asyncio.create_task(_process_document_async(d["id"], user["tenant_id"], engine_override))
            d.pop("file_b64", None)
            for p in d.get("pages") or []:
                p.pop("file_b64", None)
            saved.append(d)
        except HTTPException as e:
            saved.append({"filename": f.filename, "error": e.detail})
    return {"uploaded": len([s for s in saved if s.get("id")]), "documents": saved}

async def _process_document_async(doc_id: str, tenant_id: str, engine_override: Optional[str] = None):
    doc = await db.documents.find_one({"id": doc_id, "tenant_id": tenant_id})
    if not doc:
        return
    await db.documents.update_one({"id": doc_id}, {"$set": {"status": "processing"}})
    try:
        pages = _doc_pages(doc)
        if not pages:
            data = {"doc_type": "other", "confidence": 0.0, "raw_text": "",
                    "error": "No processable pages", "_engine": "none", "_attempts": []}
        else:
            data = await extract_with_engine_multipage(tenant_id, pages, engine_override)
        confidence = float(data.get("confidence") or 0.0)
        doc_type = data.get("doc_type") or "other"
        errors = await run_validations(tenant_id, doc_id, data)
        status = "processed" if not data.get("error") else "failed"
        await db.documents.update_one({"id": doc_id}, {"$set": {
            "extracted_data": data,
            "confidence": confidence,
            "doc_type": doc_type,
            "status": status,
            "validation_errors": errors,
            "extraction_engine": data.get("_engine"),
            "extraction_attempts": data.get("_attempts") or [],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }})
    except Exception as e:
        logger.exception("processing failed")
        await db.documents.update_one({"id": doc_id}, {"$set": {
            "status": "failed", "validation_errors": [{"field": "_system", "level": "error", "message": str(e)}]
        }})

@api.post("/documents/{doc_id}/process")
async def process_document(doc_id: str,
                           engine_override: Optional[str] = Query(None),
                           user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({**_visibility_filter(user), "id": doc_id})
    if not doc:
        raise HTTPException(404, "Not found")
    await _process_document_async(doc_id, user["tenant_id"], engine_override)
    await log_audit(user["tenant_id"], user["id"], "process_document", "document", doc_id,
                    {"engine": engine_override or "tenant_default"})
    updated = await db.documents.find_one({"id": doc_id}, {"_id": 0, "file_b64": 0, "pages.file_b64": 0})
    return updated

@api.get("/documents")
async def list_documents(user: dict = Depends(get_current_user),
                         q: Optional[str] = None,
                         doc_type: Optional[str] = None,
                         status: Optional[str] = None,
                         limit: int = 50,
                         skip: int = 0):
    query = _visibility_filter(user)
    if doc_type and doc_type != "all":
        query["doc_type"] = doc_type
    if status and status != "all":
        query["status"] = status
    if q:
        query["$or"] = [
            {"filename": {"$regex": q, "$options": "i"}},
            {"extracted_data.invoice_number": {"$regex": q, "$options": "i"}},
            {"extracted_data.vendor_name": {"$regex": q, "$options": "i"}},
            {"extracted_data.dc_number": {"$regex": q, "$options": "i"}},
            {"extracted_data.po_number": {"$regex": q, "$options": "i"}},
        ]
    total = await db.documents.count_documents(query)
    cursor = db.documents.find(query, {"_id": 0, "file_b64": 0, "pages.file_b64": 0}).sort("created_at", -1).skip(skip).limit(limit)
    docs = await cursor.to_list(limit)
    return {"total": total, "documents": docs}

@api.get("/documents/{doc_id}")
async def get_document(doc_id: str, user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({**_visibility_filter(user), "id": doc_id},
                                       {"_id": 0, "file_b64": 0, "pages.file_b64": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    return doc

@api.get("/documents/{doc_id}/file")
async def get_document_file(doc_id: str,
                            page: int = Query(1, ge=1),
                            user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({**_visibility_filter(user), "id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Not found")
    pages = _doc_pages(doc)
    if not pages:
        raise HTTPException(404, "Document has no pages")
    if page > len(pages):
        raise HTTPException(404, f"Page {page} not found (doc has {len(pages)} pages)")
    p = pages[page - 1]
    return {"filename": doc.get("filename"), "mime_type": p.get("mime_type"),
            "page_number": p.get("page_number"), "page_count": len(pages),
            "data_url": f"data:{p.get('mime_type')};base64,{p.get('file_b64')}"}

@api.put("/documents/{doc_id}")
async def update_document(doc_id: str, payload: DocumentUpdateIn,
                          user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({**_visibility_filter(user), "id": doc_id})
    if not doc:
        raise HTTPException(404, "Not found")
    update = {}
    if payload.extracted_data is not None:
        update["extracted_data"] = payload.extracted_data
        update["validation_errors"] = await run_validations(user["tenant_id"], doc_id, payload.extracted_data)
    if payload.doc_type is not None:
        update["doc_type"] = payload.doc_type
    if payload.status is not None:
        update["status"] = payload.status
    if payload.notes is not None:
        update["notes"] = payload.notes
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.documents.update_one({"id": doc_id}, {"$set": update})
    await log_audit(user["tenant_id"], user["id"], "update_document", "document", doc_id)
    return await db.documents.find_one({"id": doc_id}, {"_id": 0, "file_b64": 0, "pages.file_b64": 0})

@api.post("/documents/{doc_id}/approve")
async def approve_document(doc_id: str, user: dict = Depends(require_roles("admin", "finance", "manager"))):
    res = await db.documents.update_one(
        {"id": doc_id, "tenant_id": user["tenant_id"]},
        {"$set": {"status": "approved", "approved_by": user["id"],
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Not found")
    await log_audit(user["tenant_id"], user["id"], "approve_document", "document", doc_id)
    return await db.documents.find_one({"id": doc_id}, {"_id": 0, "file_b64": 0, "pages.file_b64": 0})

@api.post("/documents/{doc_id}/reject")
async def reject_document(doc_id: str, payload: DocumentUpdateIn,
                          user: dict = Depends(require_roles("admin", "finance", "manager"))):
    res = await db.documents.update_one(
        {"id": doc_id, "tenant_id": user["tenant_id"]},
        {"$set": {"status": "rejected", "approved_by": user["id"],
                  "notes": payload.notes or "",
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Not found")
    await log_audit(user["tenant_id"], user["id"], "reject_document", "document", doc_id)
    return await db.documents.find_one({"id": doc_id}, {"_id": 0, "file_b64": 0, "pages.file_b64": 0})

@api.delete("/documents/{doc_id}")
async def delete_document(doc_id: str, user: dict = Depends(get_current_user)):
    # admin can delete anything in their tenant; ops can delete only own; others forbidden
    if user.get("role") not in ("admin", "operations"):
        raise HTTPException(403, "Requires role: admin or operations")
    q = {"id": doc_id, "tenant_id": user["tenant_id"]}
    if user.get("role") == "operations":
        q["uploaded_by"] = user["id"]
    res = await db.documents.delete_one(q)
    if res.deleted_count == 0:
        raise HTTPException(404, "Not found")
    await log_audit(user["tenant_id"], user["id"], "delete_document", "document", doc_id)
    return {"ok": True}

# ----------------------------- COPILOT -----------------------------
@api.post("/documents/{doc_id}/copilot/chat")
async def copilot_chat(doc_id: str, payload: CopilotChatIn, user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({**_visibility_filter(user), "id": doc_id})
    if not doc:
        raise HTTPException(404, "Not found")
    pages = _doc_pages(doc)
    first_b64 = pages[0].get("file_b64", "") if pages else ""
    first_mime = pages[0].get("mime_type", "") if pages else ""
    reply = await copilot_reply(
        tenant_id=user["tenant_id"],
        doc=doc,
        file_b64=first_b64,
        mime=first_mime,
        user_msg=payload.message,
        history=payload.history or [],
    )
    await log_audit(user["tenant_id"], user["id"], "copilot_chat", "document", doc_id,
                    {"chars": len(payload.message)})
    return {"reply": reply}

# ----------------------------- EXPORTS -----------------------------
def _flatten_doc(doc: dict) -> dict:
    d = doc.get("extracted_data") or {}
    return {
        "id": doc.get("id"),
        "filename": doc.get("filename"),
        "doc_type": doc.get("doc_type"),
        "status": doc.get("status"),
        "confidence": doc.get("confidence"),
        "vendor_name": d.get("vendor_name"),
        "vendor_gstin": d.get("vendor_gstin"),
        "invoice_number": d.get("invoice_number"),
        "dc_number": d.get("dc_number"),
        "po_number": d.get("po_number"),
        "invoice_date": d.get("invoice_date"),
        "due_date": d.get("due_date"),
        "subtotal": d.get("subtotal"),
        "cgst": d.get("cgst"),
        "sgst": d.get("sgst"),
        "igst": d.get("igst"),
        "total_tax": d.get("total_tax"),
        "total_amount": d.get("total_amount"),
        "currency": d.get("currency"),
        "transport_mode": d.get("transport_mode"),
        "vehicle_number": d.get("vehicle_number"),
        "lr_number": d.get("lr_number"),
        "remarks": d.get("remarks"),
        "uploaded_by": doc.get("uploaded_by_name"),
        "created_at": doc.get("created_at"),
    }

@api.get("/documents/export/all")
async def export_all(format: str = Query("excel"), user: dict = Depends(get_current_user)):
    cursor = db.documents.find(_visibility_filter(user), {"_id": 0, "file_b64": 0, "pages.file_b64": 0})
    docs = await cursor.to_list(5000)
    rows = [_flatten_doc(d) for d in docs]
    fmt = format.lower()
    if fmt == "json":
        from json import dumps
        body = dumps({"documents": rows}, indent=2, default=str)
        return StreamingResponse(BytesIO(body.encode("utf-8")), media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=documents.json"})
    if fmt == "csv":
        import csv
        if rows:
            keys = list(rows[0].keys())
            import io
            sio = io.StringIO()
            writer = csv.DictWriter(sio, fieldnames=keys)
            writer.writeheader()
            for r in rows:
                writer.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in keys})
            data = sio.getvalue().encode("utf-8")
        else:
            data = b"no data"
        return StreamingResponse(BytesIO(data), media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=documents.csv"})
    if fmt == "xml":
        def esc(v): return ("" if v is None else str(v)).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        body = "<?xml version='1.0' encoding='UTF-8'?>\n<documents>\n"
        for r in rows:
            body += "  <document>\n"
            for k, v in r.items():
                body += f"    <{k}>{esc(v)}</{k}>\n"
            body += "  </document>\n"
        body += "</documents>"
        return StreamingResponse(BytesIO(body.encode("utf-8")), media_type="application/xml",
            headers={"Content-Disposition": "attachment; filename=documents.xml"})
    # excel default
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(k) for k in headers])
    # second sheet: line items
    li = wb.create_sheet("LineItems")
    li.append(["document_id", "vendor_name", "invoice_number", "description", "hsn_sac",
               "quantity", "unit", "unit_price", "amount", "tax_rate"])
    for d in docs:
        ed = d.get("extracted_data") or {}
        for it in ed.get("line_items") or []:
            li.append([d.get("id"), ed.get("vendor_name"), ed.get("invoice_number"),
                       it.get("description"), it.get("hsn_sac"), it.get("quantity"),
                       it.get("unit"), it.get("unit_price"), it.get("amount"), it.get("tax_rate")])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=documents.xlsx"}
    )

@api.get("/documents/{doc_id}/export")
async def export_one(doc_id: str, format: str = Query("excel"), user: dict = Depends(get_current_user)):
    d = await db.documents.find_one({**_visibility_filter(user), "id": doc_id}, {"_id": 0, "file_b64": 0, "pages.file_b64": 0})
    if not d:
        raise HTTPException(404, "Not found")
    rows = [_flatten_doc(d)]
    if format.lower() == "json":
        body = json.dumps({"document": rows[0], "raw": d.get("extracted_data")}, indent=2, default=str)
        return StreamingResponse(BytesIO(body.encode("utf-8")), media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={doc_id}.json"})
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    keys = list(rows[0].keys())
    ws.append(keys)
    ws.append([rows[0].get(k) for k in keys])
    li = wb.create_sheet("LineItems")
    li.append(["description", "hsn_sac", "quantity", "unit", "unit_price", "amount", "tax_rate"])
    ed = d.get("extracted_data") or {}
    for it in ed.get("line_items") or []:
        li.append([it.get("description"), it.get("hsn_sac"), it.get("quantity"),
                   it.get("unit"), it.get("unit_price"), it.get("amount"), it.get("tax_rate")])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={doc_id}.xlsx"})

# ----------------------------- DASHBOARD / AUDIT -----------------------------
@api.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    base = _visibility_filter(user)
    total = await db.documents.count_documents(base)
    by_status = {}
    for s in ["pending", "processing", "processed", "approved", "rejected", "failed"]:
        by_status[s] = await db.documents.count_documents({**base, "status": s})
    by_type = {}
    for t in ["invoice", "delivery_challan", "purchase_order", "grn", "packing_slip", "eway_bill", "transport_slip", "other", "unknown"]:
        c = await db.documents.count_documents({**base, "doc_type": t})
        if c > 0:
            by_type[t] = c
    pipeline = [
        {"$match": {**base, "confidence": {"$gt": 0}}},
        {"$group": {"_id": None, "avg": {"$avg": "$confidence"}}}
    ]
    avg = 0.0
    async for r in db.documents.aggregate(pipeline):
        avg = r.get("avg", 0.0) or 0.0
    today = datetime.now(timezone.utc).date()
    days = [(today - timedelta(days=i)).isoformat() for i in range(13, -1, -1)]
    trend = []
    for day in days:
        start = day + "T00:00:00"
        end = day + "T23:59:59"
        c = await db.documents.count_documents({**base, "created_at": {"$gte": start, "$lte": end}})
        trend.append({"date": day, "count": c})
    vp = [
        {"$match": {**base, "extracted_data.vendor_name": {"$nin": [None, ""]}}},
        {"$group": {"_id": "$extracted_data.vendor_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    top_vendors = []
    async for r in db.documents.aggregate(vp):
        top_vendors.append({"vendor": r["_id"], "count": r["count"]})
    pending_review = await db.documents.count_documents({**base,
        "status": {"$in": ["processed", "pending"]}})
    failed = by_status.get("failed", 0)
    return {
        "total_documents": total,
        "by_status": by_status,
        "by_type": by_type,
        "avg_confidence": round(avg, 3),
        "trend_14d": trend,
        "top_vendors": top_vendors,
        "pending_review": pending_review,
        "failed_validations": failed,
        "scope": "all" if user.get("role") in ("admin", "finance", "manager") else "own",
    }

@api.get("/audit-logs")
async def list_audit_logs(user: dict = Depends(require_roles("admin", "manager")),
                          limit: int = 100):
    cursor = db.audit_logs.find({"tenant_id": user["tenant_id"]}, {"_id": 0}).sort("created_at", -1).limit(limit)
    logs = await cursor.to_list(limit)
    # enrich with user names
    user_ids = list({log.get("user_id") for log in logs})
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(500)
    umap = {u["id"]: u for u in users}
    for log in logs:
        u = umap.get(log.get("user_id"))
        log["user_name"] = u.get("name") if u else "Unknown"
        log["user_email"] = u.get("email") if u else ""
    return logs

@api.get("/")
async def root():
    return {"name": "Document Intelligence Platform", "version": "1.0.0", "docs": "/docs"}

# ----------------------------- INCLUDE / CORS -----------------------------
app.include_router(api)

frontend_origin = os.environ.get("FRONTEND_URL", "http://localhost:3000")
allowed_origins = [frontend_origin, "http://localhost:3000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ----------------------------- STARTUP -----------------------------
@app.on_event("startup")
async def on_startup():
    await db.users.create_index("email", unique=True)
    await db.documents.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.login_attempts.create_index("identifier")
    # Seed default tenant + admin
    default = await db.tenants.find_one({"name": "Default Tenant"})
    if not default:
        tenant_id = str(uuid.uuid4())
        await db.tenants.insert_one({
            "id": tenant_id, "name": "Default Tenant", "gstin": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    else:
        tenant_id = default["id"]
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@docintel.io")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Platform Admin", "role": "admin", "tenant_id": tenant_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}})
    logger.info("Startup complete. Admin: %s", admin_email)

@app.on_event("shutdown")
async def on_shutdown():
    client.close()
